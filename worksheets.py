import datetime
import openpyxl
import configparser

import io

from sharepoint import Site
import vehicles

## Read Config files

config = configparser.ConfigParser()
config.read(['./config/authentication.ini', './config/config.ini', './config/eno.ini', './config/vehicles.ini', './config/zones.ini'])

## Set test mode & location (for excel schedule file location)

test_mode = config.getboolean('testing', 'test_mode')
print(f"Test mode is {test_mode}")

## Read the schedule into memory to prevent sharing violation in shared excel files
xlsx_filename=config['testing']['schedule']
with open(xlsx_filename, "rb") as f:
    in_mem_file = io.BytesIO(f.read())

schedule = openpyxl.load_workbook(in_mem_file, read_only=True)
print(f"Schedule location is {config['testing']['schedule']}")

## Sharepoint Authentication & Configuration

site_url = config['sharepoint']['site']
usrname = f"{config['credentials']['domain']}\\{config['credentials']['user_name']}"
passwd = config['credentials']['pass_word']

site = Site(site_url, auth=[usrname, passwd])

## Employee Numbers

trainee = []
emp_num = config['ENO']['PO'].split("\n") + config['ENO']['CSO'].split("\n") + trainee

################################################################################
# Define primary functions
################################################################################

# The main insert statment that does all the actual work
def insert_worksheet(d, s, p, e, w, z, c):
    my_data = [{'Date': d, 'WorkShift':s, 'PrecinctUnit':p, 'ENO':e, 'DutyStatus':w, 'Zone':z, 'Comments':c}]

    if test_mode:
        print(my_data)

    elif not test_mode:
        print(f"Entering {d}::{e}::{p} {z}\n")
        site.List('Daily Worksheet').UpdateListItems(data=my_data, kind='New')

def comments(w_date, section, cell_value, r_value, emp_num):

    if w_date.isoweekday() in range(0, 5):
        # Weekday Comments
        return vehicles.assign(emp_num, cell_value, r_value, w_date, config) + ' | ' + config[section][str(r_value)].split(", ")[0]
    else:
        #print(len(config[section][str(r_value)].split(", ")))
        # Weekend Comments
        if len(config[section][str(r_value)].split(", ")) == 2:
            # Check if there is a 'weekand' and a 'weekend' comment
            # If there is, return the weekend comment
            return vehicles.assign(emp_num, cell_value, r_value, w_date, config) + ' | ' + config[section][str(r_value)].split(", ")[1]
        else:
            # if not, default to the weekday comment
            return vehicles.assign(emp_num, cell_value, r_value, w_date, config) + ' | ' + config[section][str(r_value)].split(", ")[0]

def generate_eno():
    # build a dictionary of employee numbers (3 digit ENO) and sharepoint IDs.
    # The ENO is the key and the ID is the value.
    emp_id = {}
    emp_list = site.List('Employee List').GetListItems(fields=['ENO', 'Name'])

    for row in emp_list:
        try:
            emp_id[row['ENO']]=row['Name'].split('_')[0]
        except KeyError:
            pass

    return emp_id

def worksheet_helper(eno, w_date, eid, cell_value):

    for section in config.sections():
        if config.has_option(section, str(cell_value)):
            #worksheet_helper(w_date, section, cell.value, r.value)
            break

        if section == 'EOF':
            input("error")

    if section == "Admin":
        status = str(cell_value)
        zone = ""
        section = config['DEFAULT']['precinct']
        comment = config['Admin'][cell_value]

    else:
        status = 'W'
        zone = str(cell_value)
        comment = comments(w_date, section, eno, cell_value, emp_num)

    insert_worksheet(str(w_date), config['DEFAULT']['shift'], section, eid, status, zone, comment)

def rows(ws):
    shift={}
    for row in ws.iter_rows(min_row=ws.min_row, max_row=ws.max_row, min_col=2, max_col=2):
        for cell in row:
            if cell.value != None:
                shift[cell.value]=cell.row

    if config['DEFAULT']['shift'] == '1':
        row = [shift['Day Shift'],shift['Evening Shift']]
    elif config['DEFAULT']['shift'] == '2':
        row = [shift['Evening Shift'],shift['Night Shift']]
    elif config['DEFAULT']['shift'] == '3':
        row = [shift['Greek Patrol'],ws.max_row]

    return row

################################################################################
# Primary function
################################################################################
w_date = datetime.date.today()
emp_id = generate_eno()
num_days = int(input("Number of days to input: "))
#num_days = 1

for i in range(num_days):

    ws = schedule[w_date.strftime("%B")]

    eno_list = {}
    for row in ws.iter_rows(min_row=rows(ws)[0], max_row=rows(ws)[1], min_col=4, max_col=4):
        for cell in row:
            if str(cell.value) in emp_num:
                eno_list[cell.value]=[w_date, int(emp_id[str(cell.value)]), str(ws.cell(row=cell.row, column=(w_date.day+4)).value)]

    for key, value in eno_list.items():
        worksheet_helper(key, value[0], value[1], value[2])

    w_date += datetime.timedelta(1)
    print("\n")
    print((("*" * 80) + "\n") * 5)
    print("\n")

schedule._archive.close()
